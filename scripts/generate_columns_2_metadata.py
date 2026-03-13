"""
Generate COLUMNS_2.xlsx metadata workbook from CSV-driven scope.

Output layout per sheet:
  - Header row: LakehouseName | TableName | ColumnName
  - For each table:
      SOURCE marker row
      source query rows (0..N)
      TARGET marker row
      target query rows (0..N)
      3 empty rows
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from utils.fabric_client import FabricClient


DEFAULT_CSV = Path("data/etl_validation_bronze_to_silver_tests.csv")
DEFAULT_OUTPUT = Path("data/COLUMNS_2.xlsx")
HEADER = ["LakehouseName", "TableName", "ColumnName"]
TARGET_DEFAULT = "LH_Finance"
SOURCE_SCHEMA_DEFAULT = "fullload"
TARGET_SCHEMA_DEFAULT = "dbo"


def normalize_sheet_name(raw_name: str) -> str:
    name = (raw_name or "UNKNOWN").strip()
    name = re.sub(r"[:\\/?*\[\]]", "_", name)
    return name[:31] or "UNKNOWN"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate COLUMNS_2.xlsx metadata workbook.")
    parser.add_argument("--csv", dest="csv_path", default=str(DEFAULT_CSV))
    parser.add_argument("--output", dest="output_path", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--only-lakehouse", dest="only_lakehouse", default=None)
    parser.add_argument("--only-table", dest="only_table", default=None)
    return parser.parse_args()


def _csv_value(row: Dict[str, object], key: str, default: str = "") -> str:
    value = str(row.get(key, default)).strip()
    return value if value and value.upper() != "N/A" else default


def _split_csv_values(raw: object) -> List[str]:
    value = str(raw or "").strip()
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def _expand_lakehouse_pairs(row: Dict[str, object]) -> List[Tuple[str, str]]:
    source_values = _split_csv_values(row.get("source_lakehouse", ""))
    target_values = _split_csv_values(row.get("target_lakehouse", TARGET_DEFAULT))

    if not source_values:
        source_values = [""]
    if not target_values:
        target_values = [TARGET_DEFAULT]

    if len(source_values) == len(target_values):
        return list(zip(source_values, target_values))
    if len(source_values) == 1:
        return [(source_values[0], target) for target in target_values]
    if len(target_values) == 1:
        return [(source, target_values[0]) for source in source_values]

    test_id = str(row.get("test_id", "<unknown>"))
    raise ValueError(
        f"Invalid lakehouse mapping for {test_id}: source_lakehouse has {len(source_values)} values "
        f"but target_lakehouse has {len(target_values)}. Use same count or one side as a single value."
    )


def load_scope(
    csv_path: Path, only_lakehouse: str | None = None, only_table: str | None = None
) -> "OrderedDict[str, List[Tuple[str, str, str]]]":
    """
    Return OrderedDict:
      source_lakehouse -> list[(table_name, source_schema, target_lakehouse, target_schema)]
    with per-sheet table dedup and first-seen order.
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    df = pd.read_csv(csv_path).fillna("")
    df["enabled"] = df["enabled"].astype(str)
    enabled_df = df[df["enabled"].str.upper() == "TRUE"].copy()
    enabled_df["validation_type"] = enabled_df.get("validation_type", "").astype(str).str.strip().str.lower()
    enabled_rows = enabled_df[
        enabled_df["validation_type"] == "column_metadata_excel_update"
    ].to_dict("records")

    scope: "OrderedDict[str, List[Tuple[str, str, str, str]]]" = OrderedDict()
    seen_keys = set()

    only_lakehouse_norm = only_lakehouse.strip().upper() if only_lakehouse else None
    only_table_norm = only_table.strip().upper() if only_table else None

    for row in enabled_rows:
        tables_raw = _csv_value(row, "table_name")
        if not tables_raw:
            print(f"[WARN] Skipping malformed row (missing source_lakehouse/table_name): {row.get('test_id', '')}")
            continue

        source_schema = _csv_value(row, "source_schema", SOURCE_SCHEMA_DEFAULT) or SOURCE_SCHEMA_DEFAULT
        target_schema = _csv_value(row, "target_schema", TARGET_SCHEMA_DEFAULT) or TARGET_SCHEMA_DEFAULT

        tables = [t.strip() for t in tables_raw.split(",") if t.strip()]
        for source_lakehouse, target_lakehouse in _expand_lakehouse_pairs(row):
            if not source_lakehouse:
                print(f"[WARN] Skipping malformed row (missing source_lakehouse): {row.get('test_id', '')}")
                continue

            if only_lakehouse_norm and source_lakehouse.upper() != only_lakehouse_norm:
                continue

            for table_name in tables:
                if only_table_norm and table_name.upper() != only_table_norm:
                    continue

                scope.setdefault(source_lakehouse, [])
                dedup_key = (
                    source_lakehouse.upper(),
                    table_name.upper(),
                    source_schema.lower(),
                    target_lakehouse.upper(),
                    target_schema.lower(),
                )
                if dedup_key in seen_keys:
                    continue
                seen_keys.add(dedup_key)
                scope[source_lakehouse].append((table_name, source_schema, target_lakehouse, target_schema))

    return scope


def source_query(lakehouse: str, schema: str, table: str) -> str:
    return f"""
SELECT
    '{lakehouse}' AS LakehouseName,
    TABLE_NAME AS TableName,
    COLUMN_NAME AS ColumnName
FROM {lakehouse}.INFORMATION_SCHEMA.COLUMNS
WHERE TABLE_SCHEMA = '{schema}'
  AND TABLE_NAME = '{table}'
ORDER BY COLUMN_NAME;
""".strip()


def target_query(lakehouse: str, schema: str, table: str) -> str:
    return f"""
SELECT
    '{lakehouse}' AS LakehouseName,
    TABLE_NAME AS TableName,
    COLUMN_NAME AS ColumnName
FROM {lakehouse}.INFORMATION_SCHEMA.COLUMNS
WHERE TABLE_SCHEMA = '{schema}'
  AND TABLE_NAME = '{table}'
ORDER BY COLUMN_NAME;
""".strip()


def write_header(ws) -> None:
    ws.append(HEADER)
    for col, width in zip(("A", "B", "C"), (24, 36, 40)):
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")


def append_block_rows(ws, rows: Iterable[Dict[str, object]]) -> int:
    count = 0
    for row in rows:
        ws.append([
            str(row.get("LakehouseName", "")),
            str(row.get("TableName", "")),
            str(row.get("ColumnName", "")),
        ])
        count += 1
    return count


def run() -> None:
    args = parse_args()
    csv_path = Path(args.csv_path)
    output_path = Path(args.output_path)

    scope = load_scope(csv_path, args.only_lakehouse, args.only_table)
    if not scope:
        print("[INFO] No enabled rows matched filters; nothing to write.")
        return

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    source_client = FabricClient("BRONZE")
    target_client = FabricClient("SILVER")

    try:
        for source_lakehouse, table_specs in scope.items():
            ws = wb.create_sheet(title=normalize_sheet_name(source_lakehouse))
            write_header(ws)

            for table_name, src_schema, target_lakehouse, tgt_schema in table_specs:
                # SOURCE block marker
                ws.append([f"SOURCE ({source_lakehouse}.{src_schema}.{table_name})", "", ""])

                src_rows = []
                src_error = None
                try:
                    src_rows = source_client.execute_query(source_query(source_lakehouse, src_schema, table_name))
                except Exception as exc:  # noqa: BLE001
                    src_error = str(exc)

                if src_error:
                    ws.append([source_lakehouse, table_name, f"<ERROR> {src_error}"])
                    print(f"[ERROR] Source query failed: {source_lakehouse}.{src_schema}.{table_name} -> {src_error}")
                else:
                    src_count = append_block_rows(ws, src_rows)
                    if src_count == 0:
                        ws.append([source_lakehouse, table_name, "<NO_ROWS>"])

                # TARGET block marker
                ws.append([f"TARGET ({target_lakehouse}.{tgt_schema}.{table_name})", "", ""])

                tgt_rows = []
                tgt_error = None
                try:
                    tgt_rows = target_client.execute_query(target_query(target_lakehouse, tgt_schema, table_name))
                except Exception as exc:  # noqa: BLE001
                    tgt_error = str(exc)

                if tgt_error:
                    ws.append([target_lakehouse, table_name, f"<ERROR> {tgt_error}"])
                    print(f"[ERROR] Target query failed: {target_lakehouse}.{tgt_schema}.{table_name} -> {tgt_error}")
                else:
                    tgt_count = append_block_rows(ws, tgt_rows)
                    if tgt_count == 0:
                        ws.append([target_lakehouse, table_name, "<NO_ROWS>"])

                # 3 empty rows between table blocks
                ws.append(["", "", ""])
                ws.append(["", "", ""])
                ws.append(["", "", ""])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"[INFO] Metadata workbook written: {output_path}")
    finally:
        source_client.close()
        target_client.close()


if __name__ == "__main__":
    run()
