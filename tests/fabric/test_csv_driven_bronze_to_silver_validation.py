import pytest
import allure
import pandas as pd
import re
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import load_workbook
from utils.fabric_client import FabricClient
from utils.predefined_validations import PredefinedValidations

@allure.epic("ETL Testing Framework")
@allure.feature("CSV-Driven ETL Validation")
class TestCSVDrivenETLValidation:
    """CSV-Driven ETL Validation - Manual testers control tests via CSV"""
    CSV_FILE = "data/etl_validation_bronze_to_silver_tests.csv"
    SOURCE_LAYER = "BRONZE"
    TARGET_LAYER = "SILVER"
    METADATA_FILE = Path("data/COLUMNS_2.xlsx")
    _metadata_cache = None
    
    @classmethod
    def setup_class(cls):
        """Setup database clients and load CSV tests"""
        cls.source_client = FabricClient(cls.SOURCE_LAYER)
        cls.target_client = FabricClient(cls.TARGET_LAYER)
        cls.validator = PredefinedValidations()
        cls.test_cases = cls._load_test_cases()

    @classmethod
    def teardown_class(cls):
        """Close shared Fabric connections after the test class completes."""
        if getattr(cls, 'source_client', None):
            cls.source_client.close()
        if getattr(cls, 'target_client', None):
            cls.target_client.close()
    
    @classmethod
    def _load_test_cases(cls) -> List[Dict]:
        """Load test cases from CSV"""
        df = pd.read_csv(cls.CSV_FILE)
        df = df.fillna('')
        df['enabled'] = df['enabled'].astype(str)
        enabled_df = df[df['enabled'].str.upper() == 'TRUE']
        expanded_cases: List[Dict[str, Any]] = []
        for row in enabled_df.to_dict('records'):
            expanded_cases.extend(cls._expand_lakehouse_test_cases(row))
        return expanded_cases

    @staticmethod
    def _split_csv_values(raw: Any) -> List[str]:
        value = str(raw or '').strip()
        if not value:
            return []
        return [item.strip() for item in value.split(',') if item.strip()]

    @classmethod
    def _expand_lakehouse_test_cases(cls, test_case: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Expand one CSV row into one or more rows for multiple lakehouse inputs."""
        source_values = cls._split_csv_values(test_case.get('source_lakehouse', ''))
        target_values = cls._split_csv_values(test_case.get('target_lakehouse', ''))
        lhname_values = cls._split_csv_values(test_case.get('target_lhname_value', ''))

        if not source_values:
            source_values = ['']
        if not target_values:
            target_values = ['']

        if len(source_values) == len(target_values):
            pairs = list(zip(source_values, target_values))
        elif len(source_values) == 1:
            pairs = [(source_values[0], target) for target in target_values]
        elif len(target_values) == 1:
            pairs = [(source, target_values[0]) for source in source_values]
        else:
            test_id = str(test_case.get('test_id', '<unknown>'))
            raise ValueError(
                f"Invalid lakehouse mapping for {test_id}: source_lakehouse has {len(source_values)} values "
                f"but target_lakehouse has {len(target_values)}. Use same count or one side as a single value."
            )

        lhname_per_pair: List[str] = []
        if not lhname_values:
            lhname_per_pair = [src for src, _ in pairs]
        elif len(lhname_values) == len(pairs):
            lhname_per_pair = list(lhname_values)
        elif len(lhname_values) == 1 and len(pairs) > 1 and lhname_values[0] == source_values[0]:
            # Backward-compatible behavior: when CSV originally had one source and was expanded later,
            # keep lhname aligned to each expanded source lakehouse by default.
            lhname_per_pair = [src for src, _ in pairs]
        elif len(lhname_values) == 1:
            lhname_per_pair = [lhname_values[0] for _ in pairs]
        else:
            test_id = str(test_case.get('test_id', '<unknown>'))
            raise ValueError(
                f"Invalid target_lhname_value mapping for {test_id}: target_lhname_value has "
                f"{len(lhname_values)} values but expanded pairs count is {len(pairs)}. "
                "Use one value or one value per expanded pair."
            )

        if len(pairs) == 1:
            single = dict(test_case)
            single['source_lakehouse'] = pairs[0][0]
            single['target_lakehouse'] = pairs[0][1]
            single['target_lhname_value'] = lhname_per_pair[0]
            return [single]

        expanded: List[Dict[str, Any]] = []
        base_test_id = str(test_case.get('test_id', 'TEST'))
        for (source_lakehouse, target_lakehouse), lhname_value in zip(pairs, lhname_per_pair):
            row = dict(test_case)
            row['source_lakehouse'] = source_lakehouse
            row['target_lakehouse'] = target_lakehouse
            row['target_lhname_value'] = lhname_value
            row['test_id'] = f"{base_test_id}__{source_lakehouse}__{target_lakehouse}"
            expanded.append(row)
        return expanded
    
    @staticmethod
    def _resolve_query_variables(query: str, variables: Dict[str, Any]) -> str:
        """Resolve {var} placeholders in query text."""
        for var_name, var_value in variables.items():
            placeholder = f"{{{var_name}}}"
            if placeholder in query:
                if isinstance(var_value, list):
                    var_value = ','.join(map(str, var_value))
                query = query.replace(placeholder, str(var_value))
        return query

    @staticmethod
    def _execute_query_with_variables(client, query: str, variables: Dict[str, Any]) -> Any:
        """Execute query with variable substitution"""
        query = TestCSVDrivenETLValidation._resolve_query_variables(query, variables)
        return client.execute_query(query)
    
    @staticmethod
    def _extract_variables_from_query(query: str) -> List[str]:
        """Extract variable names from query"""
        return re.findall(r'\{(\w+)\}', query)

    @staticmethod
    def _derive_table_name(test_case: Dict) -> str:
        """Derive table name from CSV or source query when not explicitly provided."""
        explicit_table = str(test_case.get('table_name', '')).strip()
        if explicit_table and explicit_table.upper() != 'N/A':
            if ',' in explicit_table:
                return 'MULTI_TABLE'
            return explicit_table

        source_query = str(test_case.get('source_query', ''))
        if 'WITH Combined AS' in source_query.upper():
            return 'MULTI_TABLE'

        match = re.search(r'FROM\s+([A-Za-z0-9_\.\[\]]+)', source_query, flags=re.IGNORECASE)
        if match:
            return match.group(1)
        return 'N/A'

    @classmethod
    def _derive_lakehouse_names(cls, test_case: Dict) -> Dict[str, str]:
        """Derive source/target lakehouse names from CSV or query text."""
        source_lakehouse = cls._csv_value(test_case, 'source_lakehouse')
        target_lakehouse = cls._csv_value(test_case, 'target_lakehouse')

        source_query = str(test_case.get('source_query', ''))
        target_query = str(test_case.get('target_query', ''))

        if not source_lakehouse:
            source_match = re.search(
                r'FROM\s+([A-Za-z0-9_\[\]]+)\.[A-Za-z0-9_\[\]]+\.[A-Za-z0-9_\[\]]+',
                source_query,
                flags=re.IGNORECASE
            )
            source_lakehouse = source_match.group(1) if source_match else 'N/A'

        if not target_lakehouse:
            target_match = re.search(
                r'FROM\s+([A-Za-z0-9_\[\]]+)\.[A-Za-z0-9_\[\]]+\.[A-Za-z0-9_\[\]]+',
                target_query,
                flags=re.IGNORECASE
            )
            target_lakehouse = target_match.group(1) if target_match else 'N/A'

        return {'source_lakehouse': source_lakehouse, 'target_lakehouse': target_lakehouse}

    @staticmethod
    def _csv_value(test_case: Dict, key: str, default: str = '') -> str:
        """Read CSV field as a trimmed string with default fallback."""
        value = str(test_case.get(key, default)).strip()
        return value if value and value.upper() != 'N/A' else default

    @classmethod
    def _csv_list(cls, test_case: Dict, key: str, default: List[str]) -> List[str]:
        """Read comma-separated CSV field into list."""
        raw = cls._csv_value(test_case, key, '')
        if not raw:
            return list(default)
        return [item.strip() for item in raw.split(',') if item.strip()]

    @classmethod
    def _csv_float(cls, test_case: Dict, key: str, default: float = 0.0) -> float:
        """Read numeric CSV field with fallback."""
        raw = cls._csv_value(test_case, key, '')
        if not raw:
            return float(default)
        return float(raw)

    @classmethod
    def _build_dynamic_queries(cls, test_case: Dict) -> Dict[str, str]:
        """Build source/target queries dynamically from CSV metadata when needed.

        Expected optional CSV columns:
        - source_lakehouse, target_lakehouse
        - table_name (single table or comma-separated list)
        - source_schema (default: fullload), target_schema (default: dbo)
        - target_lhname_column (default: lhname)
        - target_lhname_value (default: source_lakehouse)
        - source_where, target_where
        """
        source_query = cls._csv_value(test_case, 'source_query') or cls._csv_value(test_case, 'source')
        target_query = cls._csv_value(test_case, 'target_query') or cls._csv_value(test_case, 'target')
        if target_query and not source_query and cls._is_target_zero_validation(
            cls._csv_value(test_case, 'validation_type')
        ):
            # Target-only validation: keep source harmless and empty-result.
            return {'source_query': 'SELECT TOP 0 1 AS dummy', 'target_query': target_query}
        if source_query and target_query:
            return {'source_query': source_query, 'target_query': target_query}

        source_lakehouse = cls._csv_value(test_case, 'source_lakehouse')
        target_lakehouse = cls._csv_value(test_case, 'target_lakehouse')
        tables_raw = cls._csv_value(test_case, 'table_name')

        if not (source_lakehouse and target_lakehouse and tables_raw):
            raise ValueError(
                "Missing query configuration. Provide source_query/target_query or "
                "source_lakehouse, target_lakehouse, and table_name."
            )

        source_schema = cls._csv_value(test_case, 'source_schema', 'fullload')
        target_schema = cls._csv_value(test_case, 'target_schema', 'dbo')
        target_lhname_column = cls._csv_value(test_case, 'target_lhname_column', 'lhname')
        target_lhname_value = cls._csv_value(test_case, 'target_lhname_value', source_lakehouse)
        source_where = cls._csv_value(test_case, 'source_where')
        target_where = cls._csv_value(test_case, 'target_where')

        tables = [table.strip() for table in tables_raw.split(',') if table.strip()]
        if not tables:
            raise ValueError("table_name must contain at least one table.")

        source_where_clause = f" WHERE {source_where}" if source_where else ""
        target_conditions = [f"{target_lhname_column}='{target_lhname_value}'"]
        if target_where:
            target_conditions.append(target_where)
        target_where_clause = " WHERE " + " AND ".join(target_conditions)

        if len(tables) == 1:
            table = tables[0]
            generated_source_query = (
                f"SELECT COUNT(*) AS row_count FROM {source_lakehouse}.{source_schema}.{table}"
                f"{source_where_clause}"
            )
            generated_target_query = (
                f"SELECT COUNT(*) AS row_count FROM {target_lakehouse}.{target_schema}.{table}"
                f"{target_where_clause}"
            )
            return {
                'source_query': source_query or generated_source_query,
                'target_query': target_query or generated_target_query,
            }

        source_parts = []
        target_parts = []
        for table in tables:
            source_parts.append(
                f"(SELECT COUNT(*) FROM {source_lakehouse}.{source_schema}.{table}{source_where_clause}) AS {table}"
            )
            target_parts.append(
                f"(SELECT COUNT(*) FROM {target_lakehouse}.{target_schema}.{table}{target_where_clause}) AS {table}"
            )

        generated_source_query = "SELECT " + ", ".join(source_parts)
        generated_target_query = "SELECT " + ", ".join(target_parts)
        return {
            'source_query': source_query or generated_source_query,
            'target_query': target_query or generated_target_query,
        }

    @classmethod
    def _build_query_variables(cls, test_case: Dict) -> Dict[str, Any]:
        """Build placeholder variables from CSV fields for query templates."""
        variables: Dict[str, Any] = {}
        for key, value in test_case.items():
            if isinstance(value, str):
                variables[key] = value.strip()
            else:
                variables[key] = value
        if not str(variables.get('target_lhname_value', '')).strip():
            variables['target_lhname_value'] = str(variables.get('source_lakehouse', '')).strip()
        return variables

    @classmethod
    def _get_table_list(cls, test_case: Dict) -> List[str]:
        """Return table names from CSV field (supports comma-separated values)."""
        raw_tables = cls._csv_value(test_case, 'table_name')
        if not raw_tables:
            return []
        return [table.strip() for table in raw_tables.split(',') if table.strip()]

    @classmethod
    def _get_dimension_list(cls, test_case: Dict) -> List[str]:
        """Return dimension names from CSV field (supports comma-separated values)."""
        raw_dimensions = cls._csv_value(test_case, 'Dimension')
        if not raw_dimensions:
            return []
        return [dimension.strip() for dimension in raw_dimensions.split(',') if dimension.strip()]

    @classmethod
    def _build_placeholder_execution_items(
        cls,
        test_case: Dict[str, Any],
        query_config: Dict[str, str],
        table_list: List[str],
        dimension_list: List[str],
    ) -> List[Dict[str, str]]:
        """Build execution items for placeholder-driven query templates."""
        template_uses_table_placeholder = (
            '{table_name}' in query_config['source_query'] or '{table_name}' in query_config['target_query']
        )
        template_uses_dimension_placeholder = (
            '{Dimension}' in query_config['source_query'] or '{Dimension}' in query_config['target_query']
        )

        if not template_uses_table_placeholder and not template_uses_dimension_placeholder:
            return []

        table_values = table_list if template_uses_table_placeholder else []
        dimension_values = dimension_list if template_uses_dimension_placeholder else []

        if template_uses_table_placeholder and not table_values:
            table_values = [cls._csv_value(test_case, 'table_name')]
        if template_uses_dimension_placeholder and not dimension_values:
            dimension_values = [cls._csv_value(test_case, 'Dimension')]

        if template_uses_table_placeholder and template_uses_dimension_placeholder:
            if len(table_values) != len(dimension_values):
                test_id = str(test_case.get('test_id', '<unknown>'))
                raise ValueError(
                    f"Invalid table/dimension pairing for {test_id}: table_name has {len(table_values)} values "
                    f"but Dimension has {len(dimension_values)}. When both placeholders are used, counts must match exactly."
                )

            return [
                {
                    'table_name': table_name,
                    'Dimension': dimension_name,
                    'label': (
                        f"{dimension_name} - {table_name}"
                        if dimension_name and table_name
                        else dimension_name or table_name or str(test_case.get('test_id', ''))
                    ),
                }
                for table_name, dimension_name in zip(table_values, dimension_values)
            ]

        if template_uses_table_placeholder:
            return [
                {
                    'table_name': table_name,
                    'Dimension': cls._csv_value(test_case, 'Dimension'),
                    'label': table_name or str(test_case.get('test_id', '')),
                }
                for table_name in table_values
            ]

        return [
            {
                'table_name': cls._csv_value(test_case, 'table_name'),
                'Dimension': dimension_name,
                'label': dimension_name or str(test_case.get('test_id', '')),
            }
            for dimension_name in dimension_values
        ]

    @staticmethod
    def _count_for_dashboard(data: Any) -> int:
        """Compute a stable count for dashboard even when validation fails."""
        try:
            df = PredefinedValidations._to_dataframe(data, dataset_name='dashboard_data')
            if len(df) == 1 and df.shape[1] >= 1:
                numeric_df = df.apply(pd.to_numeric, errors='coerce')
                if not numeric_df.isna().any().any():
                    return int(sum(numeric_df.iloc[0].tolist()))
            return int(len(df))
        except Exception:
            return 0

    @staticmethod
    def _is_recid_based_validation(validation_type: str) -> bool:
        normalized = str(validation_type).strip().lower()
        return normalized in {
            'insert_record_validation',
            'update_record_validation',
            'delete_record_validation',
            'insert_record_validation_group',
            'update_record_validation_group',
            'delete_record_validation_group',
        }

    @staticmethod
    def _is_target_zero_validation(validation_type: str) -> bool:
        normalized = str(validation_type).strip().lower()
        return normalized in {
            'target_zero_count_validation',
            'target_zero_records_validation',
            'target_should_be_zero',
        }

    @staticmethod
    def _extract_recid_list(source_results: Any) -> List[Any]:
        recids: List[Any] = []
        for row in source_results or []:
            if hasattr(row, 'get'):
                recid = row.get('recid')
                if recid is not None:
                    recids.append(recid)
        return recids

    @staticmethod
    def _count_unique_recids(rows: Any) -> int:
        """Count unique recids from row objects/dicts."""
        recids = set()
        for row in rows or []:
            if hasattr(row, 'get'):
                recid = row.get('recid')
                if recid is not None:
                    recids.add(recid)
        return len(recids)

    @staticmethod
    def _query_uses_recid_list(query: str) -> bool:
        """Check whether query template depends on {recid_list} placeholder."""
        return '{recid_list}' in str(query or '')

    @staticmethod
    def _extract_from_lakehouse(query: str) -> str:
        """Extract first lakehouse token from FROM clause when present."""
        match = re.search(
            r'FROM\s+([A-Za-z0-9_\[\]]+)\.[A-Za-z0-9_\[\]]+\.[A-Za-z0-9_\[\]]+',
            str(query or ''),
            flags=re.IGNORECASE
        )
        if not match:
            return ''
        return match.group(1).strip().strip('[]').upper()

    def _pick_client_for_query(
        self,
        query: str,
        source_lakehouse: str,
        target_lakehouse: str,
        default_side: str
    ):
        """Pick source/target Fabric client by query lakehouse reference."""
        query_lakehouse = self._extract_from_lakehouse(query)
        source_lh = str(source_lakehouse or '').strip().strip('[]').upper()
        target_lh = str(target_lakehouse or '').strip().strip('[]').upper()

        if query_lakehouse and target_lh and query_lakehouse == target_lh:
            return self.target_client
        if query_lakehouse and source_lh and query_lakehouse == source_lh:
            return self.source_client
        return self.source_client if default_side == 'source' else self.target_client

    def _execute_queries_with_dynamic_order(
        self,
        test_id: str,
        validation_type: str,
        source_query: str,
        target_query: str,
        source_lakehouse: str = '',
        target_lakehouse: str = '',
        label_suffix: str = ''
    ) -> tuple[Any, Any]:
        """Execute source/target queries honoring recid dependency direction.

        Default behavior is source -> target. If source_query depends on
        {recid_list}, execution switches to target -> source.
        """
        recid_based = self._is_recid_based_validation(validation_type)
        source_needs_recid = self._query_uses_recid_list(source_query)
        target_needs_recid = self._query_uses_recid_list(target_query)

        if source_needs_recid and target_needs_recid:
            raise AssertionError(
                "Both source_query and target_query contain {recid_list}. "
                "Only one side can depend on recid_list."
            )

        suffix = f" - {label_suffix}" if label_suffix else ""
        query_variables: Dict[str, Any] = {}
        source_query_client = self._pick_client_for_query(
            query=source_query,
            source_lakehouse=source_lakehouse,
            target_lakehouse=target_lakehouse,
            default_side='source'
        )
        target_query_client = self._pick_client_for_query(
            query=target_query,
            source_lakehouse=source_lakehouse,
            target_lakehouse=target_lakehouse,
            default_side='target'
        )

        def _lakehouse_for_client(client) -> str:
            if client is self.source_client:
                return str(source_lakehouse or '<source-unknown>')
            if client is self.target_client:
                return str(target_lakehouse or '<target-unknown>')
            return '<unknown-lakehouse>'

        def _execute_with_context(client, query_text: str, variables: Dict[str, Any], query_side: str):
            resolved_query = self._resolve_query_variables(query_text, variables)
            try:
                return client.execute_query(resolved_query)
            except Exception as exc:
                raise RuntimeError(
                    f"{query_side.capitalize()} query execution failed for test_id={test_id}{suffix}, "
                    f"lakehouse={_lakehouse_for_client(client)}, error_type={exc.__class__.__name__}, "
                    f"error={exc}, query={resolved_query}"
                ) from exc

        if source_needs_recid:
            with allure.step(f"Execute target query for {test_id}{suffix}"):
                target_results = _execute_with_context(
                    target_query_client,
                    target_query,
                    {},
                    'target'
                )
                allure.attach(
                    str(target_results[:10]),
                    name=f'Target Query Results (sample){suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )
                recid_list = self._extract_recid_list(target_results)
                query_variables['recid_list'] = recid_list
                allure.attach(
                    str(recid_list[:20]),
                    name=f'RecID List from Target (first 20){suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )
                allure.attach(
                    str(len(set(recid_list))),
                    name=f'Target RecID Count{suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )

            with allure.step(f"Execute source query for {test_id}{suffix}"):
                if recid_based and not query_variables.get('recid_list'):
                    source_results = []
                    allure.attach(
                        "Skipped source query because target returned 0 recids.",
                        name=f'Source Query Skipped{suffix}',
                        attachment_type=allure.attachment_type.TEXT
                    )
                else:
                    source_results = _execute_with_context(
                        source_query_client,
                        source_query,
                        query_variables,
                        'source'
                    )
                allure.attach(
                    str(source_results[:10]),
                    name=f'Source Query Results (sample){suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )
                if recid_based:
                    allure.attach(
                        str(self._count_unique_recids(source_results)),
                        name=f'Source RecID Count{suffix}',
                        attachment_type=allure.attachment_type.TEXT
                    )
            return source_results, target_results

        with allure.step(f"Execute source query for {test_id}{suffix}"):
            source_results = _execute_with_context(
                source_query_client,
                source_query,
                {},
                'source'
            )
            allure.attach(
                str(source_results[:10]),
                name=f'Source Query Results (sample){suffix}',
                attachment_type=allure.attachment_type.TEXT
            )
            if target_needs_recid:
                recid_list = self._extract_recid_list(source_results)
                query_variables['recid_list'] = recid_list
                allure.attach(
                    str(recid_list[:20]),
                    name=f'RecID List from Source (first 20){suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )
                allure.attach(
                    str(len(set(recid_list))),
                    name=f'Source RecID Count{suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )

        with allure.step(f"Execute target query for {test_id}{suffix}"):
            if recid_based and target_needs_recid and not query_variables.get('recid_list'):
                target_results = []
                allure.attach(
                    "Skipped target query because source returned 0 recids.",
                    name=f'Target Query Skipped{suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )
            else:
                target_results = _execute_with_context(
                    target_query_client,
                    target_query,
                    query_variables,
                    'target'
                )
            allure.attach(
                str(target_results[:10]),
                name=f'Target Query Results (sample){suffix}',
                attachment_type=allure.attachment_type.TEXT
            )
            if recid_based and target_needs_recid:
                allure.attach(
                    str(self._count_unique_recids(target_results)),
                    name=f'Target RecID Count{suffix}',
                    attachment_type=allure.attachment_type.TEXT
                )
        return source_results, target_results

    @staticmethod
    def _extract_key_set(rows: Any, key_columns: List[str]) -> set:
        """Build unique key tuples from dict-like rows for provided key columns."""
        keys = set()
        for row in rows or []:
            if not hasattr(row, 'get'):
                continue
            if all(col in row for col in key_columns):
                keys.add(tuple(row.get(col) for col in key_columns))
        return keys

    @staticmethod
    def _normalize_sheet_name(raw_name: str) -> str:
        name = (raw_name or "UNKNOWN").strip()
        name = re.sub(r"[:\\/?*\[\]]", "_", name)
        return name[:31] or "UNKNOWN"

    @classmethod
    def _load_metadata_cache(cls) -> Dict[str, Any]:
        """Load COLUMNS_2.xlsx and build lookup indexes for table columns."""
        if cls._metadata_cache is not None:
            return cls._metadata_cache

        if not cls.METADATA_FILE.exists():
            raise FileNotFoundError(f"Metadata file not found: {cls.METADATA_FILE}")

        wb = load_workbook(cls.METADATA_FILE, data_only=True)
        sheet_table_columns: Dict[tuple, List[str]] = {}
        lakehouse_table_columns: Dict[tuple, List[str]] = {}

        for ws in wb.worksheets:
            sheet_key = ws.title.strip().upper()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                lakehouse = str(row[0]).strip() if row[0] is not None else ""
                table_name = str(row[1]).strip() if row[1] is not None else ""
                column_name = str(row[2]).strip() if row[2] is not None else ""

                if (
                    not lakehouse
                    or not table_name
                    or not column_name
                    or column_name.startswith("<")
                    or lakehouse.startswith("SOURCE (")
                    or lakehouse.startswith("TARGET (")
                ):
                    continue

                table_upper = table_name.upper()
                lake_upper = lakehouse.upper()

                sheet_table_key = (sheet_key, table_upper)
                sheet_table_columns.setdefault(sheet_table_key, [])
                if column_name not in sheet_table_columns[sheet_table_key]:
                    sheet_table_columns[sheet_table_key].append(column_name)

                lake_table_key = (lake_upper, table_upper)
                lakehouse_table_columns.setdefault(lake_table_key, [])
                if column_name not in lakehouse_table_columns[lake_table_key]:
                    lakehouse_table_columns[lake_table_key].append(column_name)

        cls._metadata_cache = {
            "sheet_table_columns": sheet_table_columns,
            "lakehouse_table_columns": lakehouse_table_columns,
        }
        return cls._metadata_cache

    @classmethod
    def _get_columns_from_excel_metadata(cls, lakehouse: str, table_name: str) -> List[str]:
        """Get columns for exact lakehouse+table match from COLUMNS_2.xlsx."""
        cache = cls._load_metadata_cache()
        table_upper = table_name.strip().upper()
        lake_upper = lakehouse.strip().upper()
        sheet_upper = cls._normalize_sheet_name(lakehouse).upper()

        from_sheet = cache["sheet_table_columns"].get((sheet_upper, table_upper), [])
        if from_sheet:
            return from_sheet

        from_lakehouse = cache["lakehouse_table_columns"].get((lake_upper, table_upper), [])
        if from_lakehouse:
            return from_lakehouse

        raise AssertionError(
            f"No metadata columns found in {cls.METADATA_FILE} for lakehouse={lakehouse}, table={table_name}."
        )

    @staticmethod
    def _build_duplicate_query(lakehouse: str, schema: str, table_name: str, columns: List[str]) -> str:
        if not columns:
            raise AssertionError(f"No columns available for duplicate check on table {table_name}.")
        grouped_columns = ", ".join(f"[{col}]" for col in columns)
        return (
            f"SELECT {grouped_columns}, COUNT(*) AS duplicate_count "
            f"FROM {lakehouse}.{schema}.{table_name} "
            f"GROUP BY {grouped_columns} "
            f"HAVING COUNT(*) > 1"
        )
    
    def _execute_validation(self, test_case: Dict) -> Dict[str, Any]:
        """Execute validation based on test case configuration"""
        
        test_id = test_case['test_id']
        query_config = self._build_dynamic_queries(test_case)
        query_variables = self._build_query_variables(test_case)
        validation_type = test_case['validation_type']
        table_list = self._get_table_list(test_case)
        dimension_list = self._get_dimension_list(test_case)
        normalized_validation = str(validation_type).strip().lower()

        if normalized_validation == 'duplicate_column_check_using_excel_metadata':
            source_lakehouse = self._csv_value(test_case, 'source_lakehouse')
            target_lakehouse = self._csv_value(test_case, 'target_lakehouse', 'LH_Finance') or 'LH_Finance'
            source_schema = self._csv_value(test_case, 'source_schema', 'fullload') or 'fullload'
            target_schema = self._csv_value(test_case, 'target_schema', 'dbo') or 'dbo'

            if not table_list:
                table_name = self._derive_table_name(test_case)
                if not table_name or table_name == 'N/A':
                    raise AssertionError("table_name is required for duplicate_column_check_using_excel_metadata.")
                table_list = [table_name]

            per_table_results = []
            for table_name in table_list:
                with allure.step(f"Execute duplicate metadata validation for {test_id} - {table_name}"):
                    try:
                        source_columns = self._get_columns_from_excel_metadata(source_lakehouse, table_name)
                        target_columns = self._get_columns_from_excel_metadata(target_lakehouse, table_name)
                        source_dup_query = self._build_duplicate_query(source_lakehouse, source_schema, table_name, source_columns)
                        target_dup_query = self._build_duplicate_query(target_lakehouse, target_schema, table_name, target_columns)

                        allure.attach(source_dup_query, name=f"Source Duplicate Query - {table_name}", attachment_type=allure.attachment_type.TEXT)
                        allure.attach(target_dup_query, name=f"Target Duplicate Query - {table_name}", attachment_type=allure.attachment_type.TEXT)

                        source_duplicates = self.source_client.execute_query(source_dup_query)
                        target_duplicates = self.target_client.execute_query(target_dup_query)

                        result = self._run_validation(
                            validation_type,
                            source_duplicates,
                            target_duplicates,
                            {'table_name': table_name}
                        )
                        per_table_results.append({'table_name': table_name, 'result': result})
                    except Exception as exc:
                        per_table_results.append(
                            {
                                'table_name': table_name,
                                'result': {
                                    'status': 'ERROR',
                                    'source_count': 0,
                                    'target_count': 0,
                                    'message': (
                                        f"Metadata duplicate validation failed for table={table_name}, "
                                        f"source_lakehouse={source_lakehouse}, target_lakehouse={target_lakehouse}, "
                                        f"error_type={exc.__class__.__name__}, error={exc}"
                                    ),
                                },
                            }
                        )

            failed_tables = [item for item in per_table_results if item['result'].get('status') != 'PASSED']
            table_results = [
                {
                    'table_name': item['table_name'],
                    'status': item['result'].get('status', 'UNKNOWN'),
                    'source_count': int(item['result'].get('source_count', 0)),
                    'target_count': int(item['result'].get('target_count', 0)),
                    'message': item['result'].get('message', ''),
                }
                for item in per_table_results
            ]

            if failed_tables:
                first_failure = failed_tables[0]
                return {
                    'status': first_failure['result'].get('status', 'FAILED'),
                    'source_count': sum(int(item['result'].get('source_count', 0)) for item in per_table_results),
                    'target_count': sum(int(item['result'].get('target_count', 0)) for item in per_table_results),
                    'message': (
                        f"Duplicate metadata validation failed for {len(failed_tables)} table(s). "
                        f"First failure [{first_failure['table_name']}]: "
                        f"{first_failure['result'].get('message', 'Validation failed')}"
                    ),
                    'table_results': table_results,
                }

            return {
                'status': 'PASSED',
                'source_count': sum(int(item['result'].get('source_count', 0)) for item in per_table_results),
                'target_count': sum(int(item['result'].get('target_count', 0)) for item in per_table_results),
                'matched_count': len(per_table_results),
                'message': f"No duplicates found across {len(per_table_results)} table(s) using Excel metadata",
                'table_results': table_results,
            }

        template_uses_table_placeholder = (
            '{table_name}' in query_config['source_query'] or '{table_name}' in query_config['target_query']
        )
        template_uses_dimension_placeholder = (
            '{Dimension}' in query_config['source_query'] or '{Dimension}' in query_config['target_query']
        )

        try:
            execution_items = self._build_placeholder_execution_items(
                test_case=test_case,
                query_config=query_config,
                table_list=table_list,
                dimension_list=dimension_list,
            )
        except ValueError as exc:
            return {
                'status': 'ERROR',
                'source_count': 0,
                'target_count': 0,
                'message': str(exc),
            }

        # If query templates resolve to multiple items, execute each item independently and aggregate.
        if len(execution_items) > 1:
            per_table_results = []
            for execution_item in execution_items:
                execution_label = execution_item['label'] or test_id
                table_name = execution_item['table_name']
                dimension_name = execution_item['Dimension']

                with allure.step(f"Execute item {execution_label} for {test_id}"):
                    item_vars = dict(query_variables)
                    if template_uses_table_placeholder:
                        item_vars['table_name'] = table_name
                    if template_uses_dimension_placeholder:
                        item_vars['Dimension'] = dimension_name

                    source_query = self._resolve_query_variables(query_config['source_query'], item_vars)
                    target_query = self._resolve_query_variables(query_config['target_query'], item_vars)
                    try:
                        source_results, target_results = self._execute_queries_with_dynamic_order(
                            test_id=test_id,
                            validation_type=validation_type,
                            source_query=source_query,
                            target_query=target_query,
                            source_lakehouse=str(item_vars.get('source_lakehouse', '')),
                            target_lakehouse=str(item_vars.get('target_lakehouse', '')),
                            label_suffix=execution_label
                        )

                        with allure.step(f"Execute validation: {validation_type} - {execution_label}"):
                            runtime_test_case = dict(test_case)
                            runtime_test_case['table_name'] = table_name
                            runtime_test_case['Dimension'] = dimension_name
                            runtime_test_case['source_query'] = source_query
                            runtime_test_case['target_query'] = target_query
                            result = self._run_validation(
                                validation_type, source_results, target_results, runtime_test_case
                            )
                            per_table_results.append({'table_name': execution_label, 'result': result})
                    except Exception as exc:
                        per_table_results.append({
                            'table_name': execution_label,
                            'result': {
                                'status': 'ERROR',
                                'source_count': 0,
                                'target_count': 0,
                                'message': (
                                    f"Query execution failed for item={execution_label}, "
                                    f"source_lakehouse={item_vars.get('source_lakehouse', '')}, "
                                    f"target_lakehouse={item_vars.get('target_lakehouse', '')}, "
                                    f"error_type={exc.__class__.__name__}, error={exc}. "
                                    f"Source query: {source_query} | Target query: {target_query}"
                                )
                            }
                        })

            failed_tables = [
                item for item in per_table_results if item['result'].get('status') != 'PASSED'
            ]
            table_results = [
                {
                    'table_name': item['table_name'],
                    'status': item['result'].get('status', 'UNKNOWN'),
                    'source_count': int(item['result'].get('source_count', 0)),
                    'target_count': int(item['result'].get('target_count', 0)),
                    'source_recid_count': int(item['result'].get('source_recid_count', 0)),
                    'target_recid_count': int(item['result'].get('target_recid_count', 0)),
                    'message': item['result'].get('message', '')
                }
                                for item in per_table_results
                            ]

            if failed_tables:
                first_failure = failed_tables[0]
                return {
                    'status': first_failure['result'].get('status', 'FAILED'),
                    'source_count': sum(
                        int(item['result'].get('source_count', 0)) for item in per_table_results
                    ),
                    'target_count': sum(
                        int(item['result'].get('target_count', 0)) for item in per_table_results
                    ),
                    'message': (
                        f"Validation failed for {len(failed_tables)} item(s). "
                        f"First failure [{first_failure['table_name']}]: "
                        f"{first_failure['result'].get('message', 'Validation failed')}"
                    ),
                    'table_results': table_results
                }

            return {
                'status': 'PASSED',
                'source_count': sum(
                    int(item['result'].get('source_count', 0)) for item in per_table_results
                ),
                'target_count': sum(
                    int(item['result'].get('target_count', 0)) for item in per_table_results
                ),
                'matched_count': len(per_table_results),
                'message': f"All {len(per_table_results)} item validations passed",
                'table_results': table_results
            }

        if len(execution_items) == 1:
            single_item = execution_items[0]
            if template_uses_table_placeholder:
                query_variables['table_name'] = single_item['table_name']
            if template_uses_dimension_placeholder:
                query_variables['Dimension'] = single_item['Dimension']
        elif template_uses_dimension_placeholder and len(dimension_list) == 1:
            query_variables['Dimension'] = dimension_list[0]

        source_query = self._resolve_query_variables(query_config['source_query'], query_variables)
        target_query = self._resolve_query_variables(query_config['target_query'], query_variables)
        source_results, target_results = self._execute_queries_with_dynamic_order(
            test_id=test_id,
            validation_type=validation_type,
            source_query=source_query,
            target_query=target_query,
            source_lakehouse=str(query_variables.get('source_lakehouse', '')),
            target_lakehouse=str(query_variables.get('target_lakehouse', ''))
        )
        
        with allure.step(f"Execute validation: {validation_type}"):
            runtime_test_case = dict(test_case)
            if template_uses_table_placeholder:
                runtime_test_case['table_name'] = query_variables.get('table_name', '')
            if template_uses_dimension_placeholder:
                runtime_test_case['Dimension'] = query_variables.get('Dimension', '')
            runtime_test_case['source_query'] = source_query
            runtime_test_case['target_query'] = target_query
            result = self._run_validation(
                validation_type, source_results, target_results, runtime_test_case
            )
            return result
    
    def _run_validation(self, validation_type: str, source_data: Any, 
                       target_data: Any, test_case: Dict) -> Dict[str, Any]:
        """Run specific validation type"""
        
        try:
            normalized_validation = str(validation_type).strip().lower()

            if normalized_validation == 'row_count_comparison':
                source_count = self._count_for_dashboard(source_data)
                target_count = self._count_for_dashboard(target_data)

                source_lh = str(test_case.get('source_lakehouse', '')).upper()
                target_lh = str(test_case.get('target_lakehouse', '')).upper()
                target_lhname = str(test_case.get('target_lhname_value', '')).upper()
                is_d365_lakehouse = any(
                    'D365' in value for value in (source_lh, target_lh, target_lhname)
                )

                if is_d365_lakehouse:
                    if source_count >= target_count:
                        return {
                            'status': 'PASSED',
                            'source_count': source_count,
                            'target_count': target_count,
                            'message': (
                                f'D365 row count rule passed: source ({source_count}) >= '
                                f'target ({target_count})'
                            )
                        }
                    return {
                        'status': 'FAILED',
                        'source_count': source_count,
                        'target_count': target_count,
                        'message': (
                            f'D365 row count rule failed: source ({source_count}) must be >= '
                            f'target ({target_count})'
                        )
                    }

                try:
                    self.validator.row_count_comparison(source_data, target_data)
                except AssertionError as e:
                    return {
                        'status': 'FAILED',
                        'source_count': source_count,
                        'target_count': target_count,
                        'message': str(e)
                    }
                return {
                    'status': 'PASSED',
                    'source_count': source_count,
                    'target_count': target_count,
                    'message': f'Row counts match: {source_count}'
                }

            elif self._is_target_zero_validation(normalized_validation):
                target_rows = list(target_data or [])
                target_count = len(target_rows)
                if target_count == 0:
                    return {
                        'status': 'PASSED',
                        'source_count': len(source_data or []),
                        'target_count': 0,
                        'matched_count': 0,
                        'message': 'Target query returned 0 records as expected'
                    }

                sample_rows = target_rows[:10]
                return {
                    'status': 'FAILED',
                    'source_count': len(source_data or []),
                    'target_count': target_count,
                    'failed_count': target_count,
                    'failed_records': sample_rows,
                    'message': (
                        f"Target query returned {target_count} record(s), expected 0. "
                        f"Sample records: {sample_rows}"
                    )
                }
            
            elif normalized_validation in ('insert_record_validation', 'update_record_validation', 'delete_record_validation'):
                source_recids = set(row['recid'] for row in source_data)
                target_recids = set(row['recid'] for row in target_data)
                missing_recids = source_recids - target_recids
                
                if missing_recids:
                    return {
                        'status': 'FAILED',
                        'source_count': len(source_recids),
                        'target_count': len(target_recids),
                        'source_recid_count': len(source_recids),
                        'target_recid_count': len(target_recids),
                        'missing_count': len(missing_recids),
                        'missing_recids': list(missing_recids)[:10],
                        'message': f'{len(missing_recids)} recids missing in target'
                    }
                
                return {
                    'status': 'PASSED',
                    'source_count': len(source_recids),
                    'target_count': len(source_recids),
                    'source_recid_count': len(source_recids),
                    'target_recid_count': len(target_recids),
                    'matched_count': len(source_recids),
                    'message': (
                        'All source recids found in target'
                        if len(source_recids) > 0
                        else 'Source returned 0 recids; validation treated as PASS'
                    )
                }
            
            elif normalized_validation == 'soft_delete_consistency':
                key_columns = self._csv_list(test_case, 'key_columns', ['recid'])
                delete_column = self._csv_value(test_case, 'delete_column', 'isdelete')
                summary = self.validator.soft_delete_consistency(
                    source_data=source_data,
                    target_data=target_data,
                    key_columns=key_columns,
                    delete_column=delete_column
                )
                return {
                    'status': 'PASSED',
                    'source_count': len(source_data),
                    'target_count': len(target_data),
                    'matched_count': summary['joined_row_count'],
                    'message': f"Soft delete consistency passed for {summary['joined_row_count']} joined records"
                }

            elif normalized_validation in ('record_level_dataframe_comparison', 'record_level_comparison', 'insert_record_validation_group', 'update_record_validation_group', 'delete_record_validation_group'):
                key_columns = self._csv_list(test_case, 'key_columns', ['recid'])
                if source_data and target_data:
                    source_cols = set(source_data[0].keys()) if hasattr(source_data[0], 'keys') else set()
                    target_cols = set(target_data[0].keys()) if hasattr(target_data[0], 'keys') else set()
                    if 'TableName' in source_cols and 'TableName' in target_cols and key_columns == ['recid']:
                        key_columns = ['TableName', 'recid']

                source_query = str(test_case.get('source_query', ''))
                target_query = str(test_case.get('target_query', ''))
                if '{recid_list}' in source_query or '{recid_list}' in target_query:
                    source_keys = self._extract_key_set(source_data, key_columns)
                    target_keys = self._extract_key_set(target_data, key_columns)

                    if not source_keys:
                        return {
                            'status': 'PASSED',
                            'source_count': 0,
                            'target_count': 0,
                            'source_recid_count': 0,
                            'target_recid_count': len(target_keys),
                            'matched_count': 0,
                            'message': 'Source returned 0 recids; validation treated as PASS'
                        }

                    missing_keys = source_keys - target_keys
                    if missing_keys:
                        return {
                            'status': 'FAILED',
                            'source_count': len(source_keys),
                            'target_count': len(target_keys),
                            'source_recid_count': len(source_keys),
                            'target_recid_count': len(target_keys),
                            'missing_count': len(missing_keys),
                            'message': f'{len(missing_keys)} source keys missing in target. Sample: {list(missing_keys)[:10]}'
                        }

                    return {
                        'status': 'PASSED',
                        'source_count': len(source_keys),
                        'target_count': len(target_keys),
                        'source_recid_count': len(source_keys),
                        'target_recid_count': len(target_keys),
                        'matched_count': len(source_keys),
                        'message': f'All {len(source_keys)} source keys found in target'
                    }

                compare_columns = self._csv_list(test_case, 'compare_columns', [])
                compare_cols_arg = compare_columns if compare_columns else None

                summary = self.validator.record_level_dataframe_comparison(
                    source_data=source_data,
                    target_data=target_data,
                    key_columns=key_columns,
                    compare_columns=compare_cols_arg
                )
                return {
                    'status': 'PASSED',
                    'source_count': len(source_data),
                    'target_count': len(target_data),
                    'matched_count': len(source_data) - summary['missing_in_target_count'],
                    'message': f"Record-level comparison passed for keys {key_columns}"
                }

            elif normalized_validation in ('incremental_validation', 'incremental_delta_validation'):
                watermark_column = self._csv_value(test_case, 'watermark_column', 'dpmodifieddatetime')
                key_columns = self._csv_list(test_case, 'key_columns', ['recid'])
                delta_start = self._csv_value(test_case, 'delta_start', '')
                delta_end = self._csv_value(test_case, 'delta_end', '')
                summary = self.validator.incremental_delta_validation(
                    source_data=source_data,
                    target_data=target_data,
                    watermark_column=watermark_column,
                    key_columns=key_columns,
                    delta_start=delta_start or None,
                    delta_end=delta_end or None,
                    allow_empty_delta=True
                )
                return {
                    'status': 'PASSED',
                    'source_count': summary['source_delta_rows'],
                    'target_count': summary['target_delta_rows'],
                    'matched_count': summary['source_delta_rows'] - summary['missing_key_count'],
                    'message': f"Incremental validation passed for range {summary['delta_start']} to {summary['delta_end']}"
                }

            elif normalized_validation in ('duplicate_check', 'duplicate_checks', 'duplicate_checks_primary_keys'):
                key_columns = self._csv_list(test_case, 'key_columns', ['recid'])
                source_dups = self.validator.duplicate_checks_primary_keys(source_data, key_columns)
                target_dups = self.validator.duplicate_checks_primary_keys(target_data, key_columns)
                return {
                    'status': 'PASSED',
                    'source_count': len(source_data),
                    'target_count': len(target_data),
                    'matched_count': len(source_data),
                    'message': (
                        f"Duplicate check passed on keys {key_columns}. "
                        f"Source duplicates={source_dups}, Target duplicates={target_dups}"
                    )
                }
            elif normalized_validation == 'duplicate_column_check_using_excel_metadata':
                source_dup_groups = len(source_data or [])
                target_dup_groups = len(target_data or [])
                if source_dup_groups > 0 or target_dup_groups > 0:
                    return {
                        'status': 'FAILED',
                        'source_count': source_dup_groups,
                        'target_count': target_dup_groups,
                        'message': (
                            f"Duplicate groups found. Source={source_dup_groups}, Target={target_dup_groups}"
                        ),
                    }
                return {
                    'status': 'PASSED',
                    'source_count': 0,
                    'target_count': 0,
                    'matched_count': 0,
                    'message': 'No duplicate groups found in source and target',
                }

            elif normalized_validation in ('aggregate_validation', 'aggregate_validations'):
                aggregate_columns = self._csv_list(test_case, 'aggregate_columns', [])
                aggregate_functions = self._csv_list(test_case, 'aggregate_functions', ['sum'])
                tolerance = self._csv_float(test_case, 'aggregate_tolerance', 0.0)

                if not aggregate_columns:
                    raise AssertionError("aggregate_columns is required for aggregate_validation.")

                aggregate_config = {col: aggregate_functions for col in aggregate_columns}
                self.validator.aggregate_validations(
                    source_data=source_data,
                    target_data=target_data,
                    aggregate_config=aggregate_config,
                    tolerance=tolerance
                )
                return {
                    'status': 'PASSED',
                    'source_count': len(source_data),
                    'target_count': len(target_data),
                    'matched_count': len(source_data),
                    'message': (
                        f"Aggregate validation passed for columns {aggregate_columns} "
                        f"with functions {aggregate_functions}"
                    )
                }

            elif normalized_validation in ('referential_integrity', 'referential_integrity_validation'):
                child_fk_columns = self._csv_list(test_case, 'child_fk_columns', ['recid'])
                parent_pk_columns = self._csv_list(test_case, 'parent_pk_columns', child_fk_columns)
                missing_count = self.validator.referential_integrity_validation(
                    child_data=source_data,
                    parent_data=target_data,
                    child_fk_columns=child_fk_columns,
                    parent_pk_columns=parent_pk_columns
                )
                return {
                    'status': 'PASSED',
                    'source_count': len(source_data),
                    'target_count': len(target_data),
                    'matched_count': len(source_data) - missing_count,
                    'message': f"Referential integrity passed for FK {child_fk_columns} -> PK {parent_pk_columns}"
                }

            elif normalized_validation == "null_checks_mandatory_columns":
                mandatory_columns = self._csv_list(
                    test_case,
                    'mandatory_columns',
                    ["recid", "isdelete", "dpcreateddatetime", "dpmodifieddatetime"]
                )
                result_data = PredefinedValidations.null_checks_mandatory_columns(
                    target_data,
                    mandatory_columns
                )
                return {
                    "status": "PASSED",
                    **result_data,
                    "message": f"Validated {result_data['total_rows_verified']} rows successfully."
                }
            else:
                return {
                    'status': 'FAILED',
                    'message': f'Unknown validation type: {validation_type}'
                }
        
        except AssertionError as e:
            return {
                'status': 'FAILED',
                'source_count': self._count_for_dashboard(source_data),
                'target_count': self._count_for_dashboard(target_data),
                'message': str(e)
            }
        except Exception as e:
            return {
                'status': 'ERROR',
                'source_count': self._count_for_dashboard(source_data),
                'target_count': self._count_for_dashboard(target_data),
                'message': f'Validation error: {str(e)}'
            }
    
    def pytest_generate_tests(self, metafunc):
        """Generate tests dynamically from CSV"""
        if 'test_case' in metafunc.fixturenames:
            test_cases = self._load_test_cases()
            metafunc.parametrize('test_case', test_cases, ids=[tc['test_id'] for tc in test_cases])
    
    @pytest.mark.fabric
    @pytest.mark.etl
    def test_csv_driven_validation(self, test_case):
        """Execute CSV-driven ETL validation test"""
        
        test_id = test_case['test_id']
        test_name = test_case['test_name']
        description = test_case['description']
        table_name = self._derive_table_name(test_case)
        lakehouses = self._derive_lakehouse_names(test_case)
        validation_type = test_case['validation_type']
        
        allure.dynamic.title(f"{test_id}: {test_name}")
        allure.dynamic.description(description)
        
        allure.dynamic.label("Table", table_name)
        allure.dynamic.label("Source_Lakehouse", lakehouses['source_lakehouse'])
        allure.dynamic.label("Target_Lakehouse", lakehouses['target_lakehouse'])
        allure.dynamic.label("Validation", validation_type.replace('_', ' ').title())
        
        result = self._execute_validation(test_case)
        
        source_count = int(result.get('source_count', 0))
        target_count = int(result.get('target_count', 0))
        allure.dynamic.label("Source_Count", str(source_count))
        allure.dynamic.label("Target_Count", str(target_count))
        allure.dynamic.label("Match_Status", f"S:{source_count}=T:{target_count}")
        allure.dynamic.label("Validation_Status", str(result.get('status', 'UNKNOWN')))
        
        with allure.step("Validation Results"):
            result_summary = f"""Table: {table_name}
Source Lakehouse: {lakehouses['source_lakehouse']}
Target Lakehouse: {lakehouses['target_lakehouse']}
Validation: {validation_type}
Status: {result['status']}
Message: {result.get('message')}"""
            
            result_summary += f"\nSource Count: {source_count}"
            result_summary += f"\nTarget Count: {target_count}"
            if 'matched_count' in result:
                result_summary += f"\nMatched Records: {result['matched_count']}"
            if 'source_recid_count' in result:
                result_summary += f"\nSource RecID Count: {result['source_recid_count']}"
            if 'target_recid_count' in result:
                result_summary += f"\nTarget RecID Count: {result['target_recid_count']}"
            if result.get('table_results'):
                result_summary += "\n\nPer Table Results:"
                for table_result in result['table_results']:
                    result_summary += (
                        f"\n- {table_result.get('table_name')}: "
                        f"Status={table_result.get('status')} | "
                        f"Source={table_result.get('source_count', 0)} | "
                        f"Target={table_result.get('target_count', 0)} | "
                        f"SourceRecIDs={table_result.get('source_recid_count', 0)} | "
                        f"TargetRecIDs={table_result.get('target_recid_count', 0)} | "
                        f"Message={table_result.get('message', '')}"
                    )
            
            allure.attach(result_summary, name='Validation Summary', 
                         attachment_type=allure.attachment_type.TEXT)
        
        assert result['status'] == 'PASSED', result.get('message', 'Validation failed')
        print(f"PASS: {test_id}: {result.get('message', 'PASSED')}")
